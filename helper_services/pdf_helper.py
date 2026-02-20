import os
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.units import inch, mm
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_RIGHT
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Paragraph, Spacer, HRFlowable
from datetime import datetime, timezone

from common.common_constants import TEMP_DIR


def generate_pdf(outcome_column, causal_analysis_results, unique_id, run_ids, filters):
    # Directory of the script
    script_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    # Create a PDF document
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    pdf_filename = f"causal_explanation_report_{timestamp}-{unique_id}.pdf"
    pdf_filepath = os.path.join(TEMP_DIR, pdf_filename)
    doc = SimpleDocTemplate(
        pdf_filepath,
        pagesize=LETTER,
        rightMargin=0.75 * inch,
        leftMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch
    )

    # Create an Excel file
    xlsx_filename = f"causal_recommendations_{timestamp}-{unique_id}.xlsx"
    xlsx_filepath = os.path.join(TEMP_DIR, xlsx_filename)
    if os.path.exists(xlsx_filepath):
        os.remove(xlsx_filepath)

    # Colors
    tab_h_bg_col = colors.HexColor("#95979d")    # table header background color
    tab_h_fg_col = colors.white                    # table header foreground color
    tab_h_ln_col = colors.HexColor("#afb1b6")    # table header line color 
    accent_col = colors.HexColor("#0099cc")      # accent color
    highlt_col = "#688f22"                       # highlight color

    # Styles for text
    styles = getSampleStyleSheet()

    # Define styles
    title_style = ParagraphStyle(
        name="Title",
        parent=styles["Title"],
        alignment=TA_LEFT
    )
    subtitle_style = ParagraphStyle(
        name="Subtitle",
        parent=styles["Heading3"],
        alignment=TA_LEFT,
        spaceBefore=0,
        spaceAfter=10
    )
    header_style = ParagraphStyle(
        name="Header",
        parent=styles["Heading3"],
        alignment=TA_LEFT,
        spaceBefore=0,
        spaceAfter=10,
        textColor=accent_col
    )
    body_style = ParagraphStyle(
        name="Body",
        parent=styles["BodyText"],
        alignment=TA_JUSTIFY,
        spaceBefore=0,
        spaceAfter=10
    )
    legend_style = ParagraphStyle(
        name="Legend",
        parent=styles["BodyText"],
        alignment=TA_LEFT,
    )
    right_style = ParagraphStyle(
        name="BodyRight",
        parent=styles["BodyText"],
        alignment=TA_RIGHT
    )

    # Define elements
    spacer = Spacer(1, 10)

    separator = HRFlowable(
        width="100%",
        thickness=1,
        lineCap='butt',
        color=accent_col,
        spaceBefore=20,
        spaceAfter=20
    )

    elements = []

    # Load images using absolute paths
    img_logo_path = os.path.join(script_dir, "images", "logo", "logo.png")
    img_up_path = os.path.join(script_dir, "images", "causal_effect", "blue_up.png")
    img_down_path = os.path.join(script_dir, "images", "causal_effect", "blue_down.png")
    img_zero_path = os.path.join(script_dir, "images", "causal_effect", "zero.png")

    img_logo = Image(img_logo_path, width=52, height=52)
    img_up = Image(img_up_path, width=16, height=16)
    img_down = Image(img_down_path, width=16, height=16)
    img_zero = Image(img_zero_path, width=16, height=16)

    # Heading
    title = Paragraph("CausalBench+: Causal Explanation Report", title_style)
    subtitle = Paragraph(
        f"Generated on {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}",
        subtitle_style
    )

    heading = Table([[[title, subtitle], img_logo]], colWidths=["*", None], hAlign="LEFT", splitByRow=0)

    table_style = [
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]

    heading.setStyle(TableStyle(table_style))

    elements.append(heading)

    elements.append(separator)

    # Legend
    legend_data = [
        [img_up, f"This variable increases {outcome_column}"],
        [img_down, f"This variable decreases {outcome_column}"],
        [img_zero, f"This variable has no effect on {outcome_column}"],
    ]

    legend = Table(legend_data, hAlign="LEFT", splitByRow=0)

    legend_style = [
        # Column alignment
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (1, 0), (1, -1), "LEFT"),

        # Vertical alignment
        ('VALIGN', (0, 0), (-1, -1), "MIDDLE"),
    ]

    legend.setStyle(TableStyle(legend_style))

    elements.append(legend)

    elements.append(separator)

    if len(causal_analysis_results) == 0:
        elements.append(Paragraph(f"<b>Analysis:</b> Effects on {outcome_column} (0 experiments)", header_style))
        elements.append(Paragraph("Insufficient data to perform causal analysis.", body_style))
        elements.append(separator)
    
    else:
        # Process data
        for group, group_data in causal_analysis_results.items():
            table_data = [["Variable", "Effect", "Strength"]]
            elements.append(Paragraph(f"<b>Analysis:</b> Effects on {group} ({group_data['experiments']} experiments)", header_style))
            
            for k, v in group_data["effects"].items():
                if abs(v) > 1000 or (abs(v) < 0.01 and v != 0):
                    value_str = f"{v:.4e}"
                else:
                    value_str = f"{v:.4f}"
                    
                if v > 0:
                    table_data.append([k, img_up, value_str])
                elif v < 0:
                    table_data.append([k, img_down, value_str])
                else:
                    table_data.append([k, img_zero, value_str])
            
            # Create the table
            table = Table(table_data, colWidths="*", hAlign="LEFT", repeatRows=1)

            # Style the table
            table_style = [
                ("BACKGROUND", (0, 0), (-1, 0), tab_h_bg_col),
                ("TEXTCOLOR", (0, 0), (-1, 0), tab_h_fg_col),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

                # Column alignment
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("ALIGN", (1, 0), (1, -1), "CENTER"),
                ("ALIGN", (2, 0), (2, -1), "RIGHT"),

                # Vertical alignment
                ('VALIGN', (0, 0), (-1, -1), "MIDDLE"),

                # Outer box
                ("BOX", (0, 0), (-1, -1), 0.5, tab_h_bg_col),
            ]

            # Line below each row
            for i in range(len(table_data) - 1):
                table_style.append(("LINEBELOW", (0, i), (-1, i), 0.5, tab_h_bg_col))

            table.setStyle(TableStyle(table_style))
            
            elements.append(table)
            elements.append(spacer)

            if len(group_data["recommendations"]) > 0:
                elements.append(spacer)
                elements.append(Paragraph(f'Top 5 additional hyperparameter settings to consider to better analyse the impact on <font color="{highlt_col}">{group}</font>:', body_style))
                            
                table_data = [['Hyperparameters'] + group_data['recommend_dims']]

                for i, recommendation in enumerate(group_data["recommendations"][:5]):
                    table_data.append((f"#{i + 1}",) + recommendation[:-1])

                table_data = list(map(list, zip(*table_data)))

                for row in table_data[1:]:
                    for j in range(1, len(row)):
                        row[j] = Paragraph(f"{row[j]}", right_style)

                # Create the table
                col_widths = [None] + ["*" for i in range(len(group_data['recommend_dims']))]
                table = Table(table_data, colWidths=col_widths, hAlign="LEFT", repeatRows=1)

                # Style the table
                table_style = [
                    ("BACKGROUND", (0, 0), (-1, 0), tab_h_bg_col),
                    ("TEXTCOLOR", (0, 0), (-1, 0), tab_h_fg_col),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

                    # Column alignment
                    ("ALIGN", (0, 0), (0, -1), "LEFT"),
                    ("ALIGN", (1, 0), (-1, -1), "RIGHT"),

                    # Vertical alignment
                    ('VALIGN', (0, 0), (-1, -1), "MIDDLE"),

                    # Outer box
                    ("BOX", (0, 0), (-1, -1), 0.5, tab_h_bg_col),
                ]

                for col in range(len(table_data[0]) - 1):
                    table_style.append(("LINEAFTER", (col, 0), (col, 0), 0.5, tab_h_ln_col))
                    table_style.append(("LINEAFTER", (col, 1), (col, -1), 0.5, tab_h_bg_col))

                table.setStyle(TableStyle(table_style))

                mode = "a" if os.path.exists(xlsx_filepath) else "w"
                with pd.ExcelWriter(xlsx_filepath, engine="openpyxl", mode=mode) as writer:
                    reco_df = pd.DataFrame(group_data["recommendations"], columns=group_data['recommend_dims'] + ['Min. Dist. to Samples'])
                    reco_df.index = reco_df.index + 1
                    reco_df.to_excel(writer, sheet_name=group, index=True)

                    ws = writer.sheets[group]
                    col_idx = reco_df.columns.get_loc("Min. Dist. to Samples") + 2
                    ws.cell(row=1, column=col_idx).font = Font(color="FF0000", bold=True)

                    for column_cells in ws.columns:
                        max_length = 0
                        column = column_cells[0].column  # column index (1-based)

                        for cell in column_cells:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass

                        adjusted_width = max_length + 2  # small padding
                        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

                elements.append(table)
                elements.append(spacer)
                elements.append(Paragraph(
                    f'Please view the complete list of recommended experiments in the sheet <font color="{highlt_col}">{group}</font> of the attached Excel file <font color="{highlt_col}">{xlsx_filename}</font>.'
                    f' The column <font color="{highlt_col}">Min. Dist. to Samples</font> in the sheet specifies the minimum standardized Euclidean distance between each recommended experiment and the experiments already executed.'
                    f' The recommended hyperparameter configurations are ranked based on the inverse of the distance, thereby prioritizing configurations that are furthest from those previously considered.',
                    body_style
                ))
            
            elements.append(separator)
    
    # Runs
    elements.append(Paragraph("<b>Run IDs</b>", header_style))
    
    if len(run_ids) == 0:
        elements.append(Paragraph("No Run IDs provided.", body_style))
    else:
        run_ids = sorted(run_ids)
        run_ids = [str(run_id) for run_id in run_ids]
        elements.append(Paragraph(f"{', '.join(run_ids)}", body_style))
    
    elements.append(separator)
    
    # Filters
    elements.append(Paragraph("<b>Filters Applied</b>", header_style))

    table_data = []

    for k, v in filters.items():
        if len(v) > 0:
            table_data.append([k, Paragraph(", ".join(v))])

    if len(table_data) == 0:
        elements.append(Paragraph("No filters applied.", body_style))

    else:
        # Create the table
        filters_table = Table(table_data, colWidths=[None, "*"], hAlign="LEFT", splitByRow=0)

        # Style the table
        table_style = [
            ("BACKGROUND", (0, 0), (0, -1), tab_h_bg_col),
            ("TEXTCOLOR", (0, 0), (0, -1), tab_h_fg_col),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),

            # Column alignment
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),

            # Vertical alignment
            ('VALIGN', (0, 0), (-1, -1), "TOP"),

            # Outer box
            ("BOX", (0, 0), (-1, -1), 0.5, tab_h_bg_col),
        ]

        # Line below each row
        for i in range(len(table_data) - 1):
            table_style.append(("LINEBELOW", (0, i), (-1, i), 0.5, tab_h_bg_col))

        filters_table.setStyle(TableStyle(table_style))

        elements.append(filters_table)

    # Build the PDF
    doc.build(elements)

    return pdf_filepath, xlsx_filepath
